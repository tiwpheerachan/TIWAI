# backend/app/services/export_service.py
"""
Export Service - Enhanced Version (SunEtLune / PEAK A–U)

Key Fixes & Improvements:
1) Correct platform detection (prefer _platform_detected / _route_name) — do NOT confuse with U_group
2) Correct U_group meaning (expense group), not "META/GOOGLE"
3) Platform-aware defaults (vendor name / VAT / price_type) + validation notes
4) Resolve "ผู้รับเงิน/คู่ค้า" (D_vendor_code) into vendor CODE (Cxxxxx) via extractors/vendor_mapping.py (optional)
5) Resolve wallet code into Q_payment_method via extractors/wallet_mapping.py (optional)
6) Robust date/amount parsing + safe Excel injection prevention
7) Preserve metadata fields in rows (do not delete), but export only A–U columns
"""

from __future__ import annotations

import csv
import io
import re
import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Any, Tuple, Optional, Set

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# =========================
# PEAK Schema (A-U)
# =========================
COLUMNS: List[Tuple[str, str]] = [
    ("A_seq", "ลำดับที่*"),
    ("A_company_name", "ชื่อบริษัท"),
    ("B_doc_date", "วันที่เอกสาร"),
    ("C_reference", "อ้างอิงถึง"),
    ("D_vendor_code", "ผู้รับเงิน/คู่ค้า"),
    ("E_tax_id_13", "เลขทะเบียน 13 หลัก"),
    ("F_branch_5", "เลขสาขา 5 หลัก"),
    ("G_invoice_no", "เลขที่ใบกำกับฯ (ถ้ามี)"),
    ("H_invoice_date", "วันที่ใบกำกับฯ (ถ้ามี)"),
    ("I_tax_purchase_date", "วันที่บันทึกภาษีซื้อ (ถ้ามี)"),
    ("J_price_type", "ประเภทราคา"),
    ("K_account", "บัญชี"),
    ("L_description", "คำอธิบาย"),
    ("M_qty", "จำนวน"),
    ("N_unit_price", "ราคาต่อหน่วย"),
    ("O_vat_rate", "อัตราภาษี"),
    ("P_wht", "หัก ณ ที่จ่าย (ถ้ามี)"),
    ("Q_payment_method", "ชำระโดย"),
    ("R_paid_amount", "จำนวนเงินที่ชำระ"),
    ("S_pnd", "ภ.ง.ด. (ถ้ามี)"),
    ("T_note", "หมายเหตุ"),
    ("U_group", "กลุ่มจัดประเภท"),
]

COL_KEYS: List[str] = [k for k, _ in COLUMNS]

# =========================
# Platforms / Rules
# =========================
PLATFORMS: Set[str] = {
    "META", "GOOGLE", "SHOPEE", "LAZADA", "TIKTOK", "SPX", "THAI_TAX", "UNKNOWN"
}

# Vendor "names" (used as vendor hint/label; actual PEAK column D wants vendor code Cxxxxx in your workflow)
PLATFORM_VENDOR_NAMES = {
    "META": "Meta Platforms Ireland",
    "GOOGLE": "Google Asia Pacific",
    "SHOPEE": "Shopee",
    "LAZADA": "Lazada",
    "TIKTOK": "TikTok",
    "SPX": "Shopee Express",
    "THAI_TAX": "",   # variable
    "UNKNOWN": "",
}

# VAT rules per platform
# NOTE: VAT here is O_vat_rate (7% or NO). price type J_price_type:
# 1 = VAT separated, 2 = VAT included, 3 = no VAT
PLATFORM_VAT_RULES = {
    "META":  {"J_price_type": "3", "O_vat_rate": "NO"},
    "GOOGLE":{"J_price_type": "3", "O_vat_rate": "NO"},
    "SHOPEE":{"J_price_type": "1", "O_vat_rate": "7%"},
    "LAZADA":{"J_price_type": "1", "O_vat_rate": "7%"},
    "TIKTOK":{"J_price_type": "1", "O_vat_rate": "7%"},
    "SPX":  {"J_price_type": "1", "O_vat_rate": "7%"},
    "THAI_TAX":{"J_price_type": "1", "O_vat_rate": "7%"},
    "UNKNOWN":{"J_price_type": "1", "O_vat_rate": "7%"},
}

VAT_RATES: Set[str] = {"7%", "NO", ""}
PRICE_TYPES: Set[str] = {"1", "2", "3", ""}
PND_ALLOWED: Set[str] = {"", "1", "2", "3", "53"}

# Expense groups (your canonical set)
GROUPS: Set[str] = {
    "Marketplace Expense",
    "Advertising Expense",
    "Delivery/Logistics Expense",
    "General Expense",
    "Inventory/COGS",
    "Other Expense",
    "",
}

PLATFORM_DEFAULT_GROUP = {
    "META": "Advertising Expense",
    "GOOGLE": "Advertising Expense",
    "SHOPEE": "Marketplace Expense",
    "LAZADA": "Marketplace Expense",
    "TIKTOK": "Marketplace Expense",
    "SPX": "Delivery/Logistics Expense",
    "THAI_TAX": "General Expense",
    "UNKNOWN": "Other Expense",
}

# =========================
# Excel / CSV concerns
# =========================
TEXT_COL_KEYS: Set[str] = {
    "A_seq",
    "A_company_name",
    "C_reference",
    "D_vendor_code",
    "E_tax_id_13",
    "F_branch_5",
    "G_invoice_no",
    "J_price_type",
    "O_vat_rate",
    "S_pnd",
    "Q_payment_method",
}

NUM_COL_KEYS: Set[str] = {"M_qty", "N_unit_price", "R_paid_amount"}

DATE_COL_KEYS: Set[str] = {"B_doc_date", "H_invoice_date", "I_tax_purchase_date"}

EXCEL_INJECTION_PREFIXES = ("=", "+", "-", "@")

# Regex patterns
RE_YYYYMMDD = re.compile(r"^\d{8}$")
RE_YYYY_MM_DD = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")
RE_DD_MM_YYYY = re.compile(r"^(\d{2})/(\d{2})/(\d{4})$")
RE_YYYY_MM_DD_SLASH = re.compile(r"^(\d{4})/(\d{2})/(\d{2})$")
RE_ALL_WS = re.compile(r"\s+")
RE_AMOUNT_CLEAN = re.compile(r"[,\s]|฿|THB|บาท", re.IGNORECASE)

MAX_ROWS = 50000
MAX_CELL_LENGTH = 32767

# =========================
# Optional mappings (your project)
# =========================
try:
    # expected in your repo: backend/app/extractors/vendor_mapping.py
    from ..extractors.vendor_mapping import resolve_vendor_code  # type: ignore
except Exception:  # pragma: no cover
    resolve_vendor_code = None  # type: ignore

try:
    # expected in your repo: backend/app/extractors/wallet_mapping.py
    from ..extractors.wallet_mapping import resolve_wallet_code  # type: ignore
except Exception:  # pragma: no cover
    resolve_wallet_code = None  # type: ignore

# =========================
# Exceptions
# =========================
class ExportValidationError(Exception):
    """Raised when export data validation fails"""
    pass


class PlatformValidationError(Exception):
    """Raised when platform-specific validation fails"""
    pass


# =========================
# Helpers
# =========================
def _s(v: Any) -> str:
    if v is None:
        return ""
    try:
        s = str(v).strip()
        if len(s) > MAX_CELL_LENGTH:
            logger.warning(f"Cell value truncated (was {len(s)} chars)")
            s = s[:MAX_CELL_LENGTH]
        return s
    except Exception:
        return ""


def _escape_excel_formula(s: str) -> str:
    if not s:
        return s
    try:
        return ("'" + s) if s[:1] in EXCEL_INJECTION_PREFIXES else s
    except Exception:
        return s


def _compact_no_ws(v: Any) -> str:
    s = _s(v)
    if not s:
        return ""
    try:
        return RE_ALL_WS.sub("", s)
    except Exception:
        return s


def _to_tax13(v: Any) -> str:
    s = _s(v)
    if not s:
        return ""
    digits = "".join(c for c in s if c.isdigit())
    return digits[:13] if len(digits) >= 13 else ""


def _to_branch5(v: Any) -> str:
    s = _s(v)
    if not s:
        return "00000"
    digits = "".join(c for c in s if c.isdigit())
    if not digits:
        return "00000"
    return digits.zfill(5)[:5]


def _parse_date_to_yyyymmdd(date_str: Any) -> str:
    if date_str is None:
        return ""
    s = _s(date_str)
    if not s:
        return ""
    try:
        if RE_YYYYMMDD.match(s):
            return s

        m = RE_YYYY_MM_DD.match(s)
        if m:
            return f"{m.group(1)}{m.group(2)}{m.group(3)}"

        m = RE_YYYY_MM_DD_SLASH.match(s)
        if m:
            return f"{m.group(1)}{m.group(2)}{m.group(3)}"

        m = RE_DD_MM_YYYY.match(s)
        if m:
            yyyy = int(m.group(3))
            # Handle Buddhist year (>= 2400) → Gregorian
            if yyyy >= 2400:
                yyyy -= 543
            return f"{yyyy:04d}{m.group(2)}{m.group(1)}"

        # Try common formats
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime("%Y%m%d")
            except ValueError:
                continue

        return ""
    except Exception:
        return ""


def _parse_amount(amount_str: Any) -> str:
    if amount_str is None:
        return ""
    s = _s(amount_str)
    if not s:
        return ""
    try:
        # keep minus? For your use case: expenses are positive; if negative, return ""
        cleaned = RE_AMOUNT_CLEAN.sub("", s).strip()
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]
        cleaned = cleaned.replace("−", "-")
        d = Decimal(cleaned)
        if d < 0:
            return ""
        return f"{d:.2f}"
    except (InvalidOperation, ValueError):
        return ""


def _clamp_choice(v: Any, allowed: Set[str], fallback: str) -> str:
    s = _s(v)
    return s if s in allowed else fallback


def _append_note(rr: Dict[str, Any], msg: str, max_lines: int = 6) -> None:
    msg = _s(msg)
    if not msg:
        return
    cur = _s(rr.get("T_note"))
    lines = [ln.strip() for ln in cur.splitlines() if ln.strip()] if cur else []
    if msg not in lines:
        lines.append(msg)
    rr["T_note"] = "\n".join(lines[:max_lines])[:1500]


# =========================
# Platform detection
# =========================
def _detect_platform(row: Dict[str, Any]) -> str:
    """
    Detect platform from the row.
    Priority:
    1) _platform_detected (from AI)
    2) _route_name
    3) vendor name hint (if D_vendor_code currently contains vendor name)
    4) group hint (if U_group looks like a group, do NOT treat as platform)
    """
    try:
        p = _s(row.get("_platform_detected")).upper()
        if p in PLATFORMS:
            return p

        route = _s(row.get("_route_name")).lower()
        if "meta" in route:
            return "META"
        if "google" in route:
            return "GOOGLE"
        if "spx" in route or "shopee_express" in route or "shopee-express" in route:
            return "SPX"
        if "shopee" in route:
            return "SHOPEE"
        if "lazada" in route:
            return "LAZADA"
        if "tiktok" in route:
            return "TIKTOK"
        if "thai" in route and "tax" in route:
            return "THAI_TAX"

        vendor = _s(row.get("D_vendor_code")).lower()
        # if already a Cxxxxx code → can't infer platform from it reliably
        if vendor.startswith("c") and vendor[1:].isdigit():
            # try fallback from other fields
            pass
        else:
            if "meta" in vendor or "facebook" in vendor or "instagram" in vendor:
                return "META"
            if "google" in vendor:
                return "GOOGLE"
            if "shopee express" in vendor or "spx" in vendor:
                return "SPX"
            if "shopee" in vendor or "ช้อปปี้" in vendor or "ช็อปปี้" in vendor:
                return "SHOPEE"
            if "lazada" in vendor or "ลาซาด้า" in vendor:
                return "LAZADA"
            if "tiktok" in vendor or "ติ๊กต๊อก" in vendor:
                return "TIKTOK"

        # If looks like Thai tax invoice (has tax id 13 + invoice no)
        tax13 = _to_tax13(row.get("E_tax_id_13"))
        inv = _s(row.get("G_invoice_no"))
        if tax13 and inv:
            return "THAI_TAX"

        return "UNKNOWN"
    except Exception:
        return "UNKNOWN"


# =========================
# Platform enforcement
# =========================
def _enforce_platform_rules(rr: Dict[str, Any], platform: str) -> None:
    # VAT rules
    rules = PLATFORM_VAT_RULES.get(platform) or PLATFORM_VAT_RULES["UNKNOWN"]
    rr["J_price_type"] = _clamp_choice(rr.get("J_price_type"), PRICE_TYPES, rules["J_price_type"])
    rr["O_vat_rate"] = _clamp_choice(rr.get("O_vat_rate"), VAT_RATES, rules["O_vat_rate"])

    # Default group (ensure it is a group, not platform code)
    ug = _s(rr.get("U_group"))
    if ug not in GROUPS:
        rr["U_group"] = PLATFORM_DEFAULT_GROUP.get(platform, "Other Expense")
    elif not ug:
        rr["U_group"] = PLATFORM_DEFAULT_GROUP.get(platform, "Other Expense")

    # Vendor name hint (kept as hint; actual PEAK D field should become vendor code Cxxxxx via mapping if possible)
    # We'll only set vendor name if empty AND mapping later fails.
    if not _s(rr.get("D_vendor_code")):
        rr["D_vendor_code"] = PLATFORM_VENDOR_NAMES.get(platform, "") or ""


# =========================
# Vendor code mapping (Cxxxxx)
# =========================
def _maybe_resolve_vendor_code(rr: Dict[str, Any]) -> None:
    """
    Convert rr["D_vendor_code"] into PEAK vendor CODE (Cxxxxx) when possible,
    based on your vendor_mapping.py (client_tax_id + vendor tax id/name).
    Safe: if resolver not available or can't resolve, keep existing value.
    """
    if resolve_vendor_code is None:
        return

    # Your workflow: use client tax id to decide mapping table
    client_tax_id = _to_tax13(rr.get("_client_tax_id") or rr.get("client_tax_id") or rr.get("A_company_tax_id"))
    vendor_tax_id = _to_tax13(rr.get("E_tax_id_13"))
    vendor_name = _s(rr.get("_vendor_name_hint") or rr.get("D_vendor_code"))

    try:
        code = resolve_vendor_code(
            client_tax_id=client_tax_id,
            vendor_tax_id=vendor_tax_id,
            vendor_name=vendor_name,
        )
        if code:
            rr["D_vendor_code"] = _s(code)
            return
    except Exception as e:
        logger.warning(f"resolve_vendor_code error: {e}")

    # If cannot resolve, keep existing D_vendor_code as-is


# =========================
# Wallet code mapping (EWLxxx)
# =========================
def _maybe_resolve_wallet(rr: Dict[str, Any]) -> None:
    """
    If wallet_mapping.py exists, set Q_payment_method to wallet code
    using seller_id/shop_name/ocr hints stored in metadata.
    Safe fallback: keep existing.
    """
    if resolve_wallet_code is None:
        return

    # Prefer metadata from extract step
    seller_id = _s(rr.get("_seller_id") or rr.get("seller_id"))
    shop_name = _s(rr.get("_shop_name") or rr.get("shop_name"))
    ocr_text_hint = _s(rr.get("_ocr_text_hint") or "")

    # If already set and looks like EWLxxx, keep it
    cur = _s(rr.get("Q_payment_method"))
    if cur.upper().startswith("EWL") and cur[3:].isdigit():
        return

    try:
        code = resolve_wallet_code(
            seller_id=seller_id,
            shop_name=shop_name,
            text=ocr_text_hint,
        )
        if code:
            rr["Q_payment_method"] = _s(code)
    except Exception as e:
        logger.warning(f"resolve_wallet_code error: {e}")


# =========================
# Validation notes
# =========================
def _platform_validation_notes(rr: Dict[str, Any], platform: str) -> None:
    # lightweight: do not block export, just add warnings
    if platform in {"META", "GOOGLE"}:
        if _s(rr.get("O_vat_rate")).upper() != "NO":
            _append_note(rr, f"⚠️ {platform}: VAT ควรเป็น NO")
        if not _s(rr.get("C_reference")):
            _append_note(rr, f"⚠️ {platform}: ไม่มีเลขอ้างอิง (C_reference)")
        if not _s(rr.get("R_paid_amount")):
            _append_note(rr, f"⚠️ {platform}: ไม่มีจำนวนเงินชำระ (R_paid_amount)")

    if platform == "THAI_TAX":
        if len(_to_tax13(rr.get("E_tax_id_13"))) != 13:
            _append_note(rr, "⚠️ THAI_TAX: ไม่พบเลขผู้เสียภาษี 13 หลักของผู้ขาย")
        if len(_to_branch5(rr.get("F_branch_5"))) != 5:
            _append_note(rr, "⚠️ THAI_TAX: เลขสาขาควรเป็น 5 หลัก (00000 ได้)")
        if not _s(rr.get("G_invoice_no")):
            _append_note(rr, "⚠️ THAI_TAX: ควรมีเลขที่ใบกำกับฯ")


# =========================
# Preprocess
# =========================
def _preprocess_rows_for_export(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    seq = 1

    logger.info(f"Preprocessing {len(rows)} rows...")

    for idx, r in enumerate(rows or [], start=1):
        try:
            rr = dict(r or {})

            # 1) detect platform
            platform = _detect_platform(rr)
            rr["_platform_final"] = platform  # keep metadata for debugging

            # 2) enforce platform rules
            _enforce_platform_rules(rr, platform)

            # 3) sequence
            rr["A_seq"] = str(seq)
            seq += 1

            # 4) policy: WHT blank (as your current policy)
            rr["P_wht"] = ""

            # 5) normalize tax/branch
            rr["E_tax_id_13"] = _to_tax13(rr.get("E_tax_id_13"))
            rr["F_branch_5"] = _to_branch5(rr.get("F_branch_5"))

            # 6) parse dates
            for dk in DATE_COL_KEYS:
                rr[dk] = _parse_date_to_yyyymmdd(rr.get(dk))

            # 7) parse amounts
            # qty: allow integer-like
            qty = _parse_amount(rr.get("M_qty"))
            rr["M_qty"] = (str(int(float(qty))) if qty and abs(float(qty) - int(float(qty))) < 1e-9 else (_s(rr.get("M_qty")) or "1"))
            rr["N_unit_price"] = _parse_amount(rr.get("N_unit_price"))
            rr["R_paid_amount"] = _parse_amount(rr.get("R_paid_amount"))

            # fallback: paid_amount from unit_price if missing
            if not rr["R_paid_amount"] and rr["N_unit_price"]:
                rr["R_paid_amount"] = rr["N_unit_price"]

            # 8) compact ids
            rr["C_reference"] = _compact_no_ws(rr.get("C_reference"))
            rr["G_invoice_no"] = _compact_no_ws(rr.get("G_invoice_no"))

            # 9) ensure strings and not None for critical
            rr["A_company_name"] = _s(rr.get("A_company_name"))
            for key in ("D_vendor_code", "E_tax_id_13", "F_branch_5", "J_price_type", "O_vat_rate", "S_pnd", "Q_payment_method"):
                if rr.get(key) is None:
                    rr[key] = ""

            # 10) clamp vat / price_type / pnd
            rr["O_vat_rate"] = _clamp_choice(rr.get("O_vat_rate"), VAT_RATES, PLATFORM_VAT_RULES.get(platform, PLATFORM_VAT_RULES["UNKNOWN"])["O_vat_rate"])
            rr["J_price_type"] = _clamp_choice(rr.get("J_price_type"), PRICE_TYPES, PLATFORM_VAT_RULES.get(platform, PLATFORM_VAT_RULES["UNKNOWN"])["J_price_type"])
            rr["S_pnd"] = _clamp_choice(rr.get("S_pnd"), PND_ALLOWED, "")

            # 11) mapping: vendor code Cxxxxx (สำคัญ)
            _maybe_resolve_vendor_code(rr)

            # 12) mapping: wallet code EWLxxx (ถ้ามี)
            _maybe_resolve_wallet(rr)

            # 13) validate notes (non-blocking)
            _platform_validation_notes(rr, platform)

            # 14) keep note clean + prevent injection
            rr["T_note"] = _escape_excel_formula(_s(rr.get("T_note")))

            # 15) If description empty: use group
            if not _s(rr.get("L_description")):
                rr["L_description"] = rr.get("U_group") or PLATFORM_DEFAULT_GROUP.get(platform, "Other Expense")

            out.append(rr)

        except Exception as e:
            logger.error(f"Error preprocessing row {idx}: {e}", exc_info=True)
            continue

    logger.info(f"Preprocessing complete: {len(out)} rows ready")
    return out


# =========================
# Validation
# =========================
def validate_rows(rows: List[Dict[str, Any]]) -> Tuple[bool, List[str]]:
    errors: List[str] = []

    if not rows:
        return (False, ["No rows to export"])

    if len(rows) > MAX_ROWS:
        return (False, [f"Too many rows: {len(rows)} (max: {MAX_ROWS})"])

    # check basic structure in first 100 rows
    for idx, row in enumerate(rows[:100], start=1):
        if not isinstance(row, dict):
            errors.append(f"Row {idx}: Not a dictionary")
            continue
        if not any(_s(row.get(k)) for k in COL_KEYS):
            errors.append(f"Row {idx}: All fields empty")

    return (len(errors) == 0, errors)


# =========================
# Type Conversion
# =========================
def _to_number_or_text(key: str, raw: Any) -> Tuple[Any, str]:
    try:
        s = _s(raw)

        # Always text
        if key in TEXT_COL_KEYS or key in DATE_COL_KEYS:
            return (_escape_excel_formula(s), numbers.FORMAT_TEXT)

        # Numeric fields
        if key in NUM_COL_KEYS:
            if key == "M_qty":
                parsed = _parse_amount(s)
                if parsed:
                    try:
                        f = float(parsed)
                        if abs(f - int(f)) < 1e-9:
                            return (int(f), "0")
                        return (f, numbers.FORMAT_NUMBER_00)
                    except Exception:
                        pass
                return (_escape_excel_formula(s), numbers.FORMAT_TEXT)

            parsed = _parse_amount(s)
            if parsed:
                try:
                    return (float(parsed), numbers.FORMAT_NUMBER_00)
                except Exception:
                    pass
            return (_escape_excel_formula(s), numbers.FORMAT_TEXT)

        return (_escape_excel_formula(s), numbers.FORMAT_TEXT)
    except Exception:
        return ("", numbers.FORMAT_TEXT)


# =========================
# Auto-fit Columns
# =========================
def _auto_fit_columns(ws, max_width: int = 60, min_width: int = 10) -> None:
    try:
        for col_idx, (_key, label) in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(label))

            max_check = min(ws.max_row + 1, 102)
            for row_idx in range(2, max_check):
                try:
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v is None:
                        continue
                    s = str(v)
                    if "\n" in s:
                        s = s.split("\n", 1)[0]
                    max_len = max(max_len, len(s))
                except Exception:
                    continue

            ws.column_dimensions[col_letter].width = int(min(max(max_len + 2, min_width), max_width))
    except Exception as e:
        logger.error(f"Auto-fit columns error: {e}")


# =========================
# CSV Export
# =========================
def export_rows_to_csv_bytes(rows: List[Dict[str, Any]]) -> bytes:
    try:
        logger.info(f"Starting CSV export for {len(rows)} rows")

        is_valid, errors = validate_rows(rows)
        if not is_valid:
            raise ExportValidationError("; ".join(errors))

        rows2 = _preprocess_rows_for_export(rows)
        if not rows2:
            raise ExportValidationError("No valid rows after preprocessing")

        out = io.StringIO()
        wri = csv.writer(out, quoting=csv.QUOTE_MINIMAL)

        wri.writerow([label for _k, label in COLUMNS])

        for r in rows2:
            row_out: List[str] = []
            for k, _label in COLUMNS:
                s = _escape_excel_formula(_s(r.get(k, "")))
                row_out.append(s)
            wri.writerow(row_out)

        result = out.getvalue().encode("utf-8-sig")
        logger.info(f"✅ CSV export complete: {len(result)} bytes")
        return result

    except ExportValidationError:
        raise
    except Exception as e:
        logger.error(f"CSV export error: {e}", exc_info=True)
        raise Exception(f"CSV export failed: {str(e)}")


# =========================
# XLSX Export
# =========================
def export_rows_to_xlsx_bytes(rows: List[Dict[str, Any]]) -> bytes:
    try:
        logger.info(f"Starting XLSX export for {len(rows)} rows")

        is_valid, errors = validate_rows(rows)
        if not is_valid:
            raise ExportValidationError("; ".join(errors))

        rows2 = _preprocess_rows_for_export(rows)
        if not rows2:
            raise ExportValidationError("No valid rows after preprocessing")

        wb = Workbook()
        ws = wb.active
        ws.title = "PEAK_IMPORT"

        headers = [label for _k, label in COLUMNS]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="E8F1FF")
        header_font = Font(bold=True)
        header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)

        thin = Side(style="thin", color="D0D7E2")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx in range(1, len(COLUMNS) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = header_align
            c.border = border

        for _row_num, r in enumerate(rows2, start=2):
            values: List[Any] = []
            formats: List[str] = []

            for k, _label in COLUMNS:
                v, fmt = _to_number_or_text(k, r.get(k, ""))
                values.append(v)
                formats.append(fmt)

            ws.append(values)

            current_row = ws.max_row
            for col_idx, fmt in enumerate(formats, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                if fmt:
                    cell.number_format = fmt
                cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in {13, 21}))
                cell.border = border

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

        col_index = {k: i + 1 for i, (k, _) in enumerate(COLUMNS)}
        for key in (TEXT_COL_KEYS | DATE_COL_KEYS):
            ci = col_index.get(key)
            if not ci:
                continue
            for row_i in range(2, 2 + len(rows2)):
                try:
                    ws.cell(row=row_i, column=ci).number_format = numbers.FORMAT_TEXT
                except Exception:
                    continue

        _auto_fit_columns(ws)

        bio = io.BytesIO()
        wb.save(bio)
        result = bio.getvalue()
        logger.info(f"✅ XLSX export complete: {len(result)} bytes")
        return result

    except ExportValidationError:
        raise
    except Exception as e:
        logger.error(f"XLSX export error: {e}", exc_info=True)
        raise Exception(f"XLSX export failed: {str(e)}")


# =========================
# Summary
# =========================
def get_export_summary(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    try:
        summary = {
            "total_rows": len(rows),
            "valid_rows": 0,
            "platforms": {},
            "extraction_methods": {},
            "clients": {},
            "date_range": {"earliest": None, "latest": None},
            "total_amount": 0.0,
            "warnings": [],
        }

        for row in rows or []:
            if _s(row.get("D_vendor_code")):
                summary["valid_rows"] += 1

            platform = _detect_platform(row)
            summary["platforms"][platform] = summary["platforms"].get(platform, 0) + 1

            method = _s(row.get("_extraction_method")) or "unknown"
            summary["extraction_methods"][method] = summary["extraction_methods"].get(method, 0) + 1

            company = _s(row.get("A_company_name")) or "Unknown"
            summary["clients"][company] = summary["clients"].get(company, 0) + 1

            doc_date = _parse_date_to_yyyymmdd(row.get("B_doc_date"))
            if doc_date:
                if not summary["date_range"]["earliest"] or doc_date < summary["date_range"]["earliest"]:
                    summary["date_range"]["earliest"] = doc_date
                if not summary["date_range"]["latest"] or doc_date > summary["date_range"]["latest"]:
                    summary["date_range"]["latest"] = doc_date

            amt = _parse_amount(row.get("R_paid_amount"))
            if amt:
                try:
                    summary["total_amount"] += float(amt)
                except Exception:
                    pass

        return summary
    except Exception as e:
        logger.error(f"Error getting export summary: {e}")
        return {"error": str(e)}


__all__ = [
    "COLUMNS",
    "export_rows_to_csv_bytes",
    "export_rows_to_xlsx_bytes",
    "ExportValidationError",
    "PlatformValidationError",
    "validate_rows",
    "get_export_summary",
]
