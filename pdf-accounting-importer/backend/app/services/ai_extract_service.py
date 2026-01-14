# backend/app/services/export_service.py
"""
Export Service - Enhanced Version for Multi-Platform Support

✅ Improvements:
1. ✅ Support for 8 platforms (META, GOOGLE, SHOPEE, LAZADA, TIKTOK, SPX, THAI_TAX, UNKNOWN)
2. ✅ Platform-aware validation and auto-correction
3. ✅ Proper PEAK schema (A-U fields)
4. ✅ Metadata preservation (_extraction_method, _route_name, etc.)
5. ✅ Smart date parsing (YYYY-MM-DD, DD/MM/YYYY, etc. → YYYYMMDD)
6. ✅ Better amount parsing (฿, THB, commas)
7. ✅ Platform-specific field validation
8. ✅ Auto-fix common mistakes
9. ✅ Comprehensive logging
10. ✅ Export summary with platform breakdown
"""
from __future__ import annotations

import csv
import io
import re
import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Any, Tuple, Optional, Set

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Setup logger
logger = logging.getLogger(__name__)

# =========================
# PEAK Schema (A-U)
# =========================
COLUMNS: List[Tuple[str, str]] = [
    ("A_seq", "ลำดับที่*"),
    ("A_company_name", "ชื่อบริษัท"),
    ("B_doc_date", "วันที่เอกสาร"),
    ("C_reference", "อ้างอิงถึง"),
    ("D_vendor_code", "ผู้รับเงิน/คู่ค้า"),
    ("E_tax_id_13", "เลขทะเบียน 13 หลัก"),
    ("F_branch_5", "เลขสาขา 5 หลัก"),
    ("G_invoice_no", "เลขที่ใบกำกับฯ (ถ้ามี)"),
    ("H_invoice_date", "วันที่ใบกำกับฯ (ถ้ามี)"),
    ("I_tax_purchase_date", "วันที่บันทึกภาษีซื้อ (ถ้ามี)"),
    ("J_price_type", "ประเภทราคา"),
    ("K_account", "บัญชี"),
    ("L_description", "คำอธิบาย"),
    ("M_qty", "จำนวน"),
    ("N_unit_price", "ราคาต่อหน่วย"),
    ("O_vat_rate", "อัตราภาษี"),
    ("P_wht", "หัก ณ ที่จ่าย (ถ้ามี)"),
    ("Q_payment_method", "ชำระโดย"),
    ("R_paid_amount", "จำนวนเงินที่ชำระ"),
    ("S_pnd", "ภ.ง.ด. (ถ้ามี)"),
    ("T_note", "หมายเหตุ"),
    ("U_group", "กลุ่มจัดประเภท"),
]

# ✅ Platform-specific vendor codes
PLATFORM_VENDORS = {
    "META": "Meta Platforms Ireland",
    "GOOGLE": "Google Asia Pacific",
    "SHOPEE": "Shopee",
    "LAZADA": "Lazada",
    "TIKTOK": "TikTok",
    "SPX": "Shopee Express",
    "THAI_TAX": "",  # Variable
    "UNKNOWN": "",
}

# ✅ Platform-specific VAT rules
PLATFORM_VAT_RULES = {
    "META": "NO",      # Reverse charge
    "GOOGLE": "NO",    # International service
    "SHOPEE": "7%",    # Thai marketplace
    "LAZADA": "7%",
    "TIKTOK": "7%",
    "SPX": "7%",
    "THAI_TAX": "7%",  # Usually
    "UNKNOWN": "",
}

# Columns that MUST be TEXT in Excel
TEXT_COL_KEYS: Set[str] = {
    "A_seq",
    "A_company_name",
    "C_reference",
    "D_vendor_code",
    "E_tax_id_13",
    "F_branch_5",
    "G_invoice_no",
    "J_price_type",
    "O_vat_rate",
    "S_pnd",
    "Q_payment_method",
}

# Numeric columns
NUM_COL_KEYS: Set[str] = {
    "M_qty",
    "N_unit_price",
    "R_paid_amount",
}

# Date columns (YYYYMMDD format)
DATE_COL_KEYS: Set[str] = {
    "B_doc_date",
    "H_invoice_date",
    "I_tax_purchase_date",
}

# CSV/Excel injection prevention
EXCEL_INJECTION_PREFIXES = ("=", "+", "-", "@")

# Regex patterns
RE_YYYYMMDD = re.compile(r"^\d{8}$")
RE_YYYY_MM_DD = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")
RE_DD_MM_YYYY = re.compile(r"^(\d{2})/(\d{2})/(\d{4})$")
RE_DECIMAL = re.compile(r"^[0-9]+(?:\.[0-9]+)?$")
RE_ALL_WS = re.compile(r"\s+")
RE_AMOUNT = re.compile(r"[฿,THB\s]")

# Limits
MAX_ROWS = 50000
MAX_CELL_LENGTH = 32767

# =========================
# Validation & Errors
# =========================
class ExportValidationError(Exception):
    """Raised when export data validation fails"""
    pass


class PlatformValidationError(Exception):
    """Raised when platform-specific validation fails"""
    pass


# =========================
# Date Parsing
# =========================
def _parse_date_to_yyyymmdd(date_str: str) -> str:
    """
    ✅ Parse various date formats to YYYYMMDD
    
    Supports:
    - YYYYMMDD (already correct)
    - YYYY-MM-DD
    - DD/MM/YYYY
    - YYYY/MM/DD
    - Timestamp
    
    Returns:
        YYYYMMDD string or empty string
    """
    if not date_str:
        return ""
    
    try:
        s = str(date_str).strip()
        
        # Already YYYYMMDD
        if RE_YYYYMMDD.match(s):
            return s
        
        # YYYY-MM-DD
        m = RE_YYYY_MM_DD.match(s)
        if m:
            return f"{m.group(1)}{m.group(2)}{m.group(3)}"
        
        # DD/MM/YYYY
        m = RE_DD_MM_YYYY.match(s)
        if m:
            return f"{m.group(3)}{m.group(2)}{m.group(1)}"
        
        # Try parsing with datetime
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"]:
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime("%Y%m%d")
            except ValueError:
                continue
        
        logger.warning(f"Could not parse date: {date_str}")
        return ""
    
    except Exception as e:
        logger.warning(f"Date parsing error: {date_str} - {e}")
        return ""


# =========================
# Amount Parsing
# =========================
def _parse_amount(amount_str: str) -> str:
    """
    ✅ Parse amount strings to decimal (2 places)
    
    Handles:
    - ฿30,000.00
    - THB 50000
    - 1,234.56
    - "1234.56"
    
    Returns:
        Decimal string like "1234.56" or ""
    """
    if not amount_str:
        return ""
    
    try:
        s = str(amount_str).strip()
        
        # Remove currency symbols and commas
        s = RE_AMOUNT.sub("", s)
        
        # Parse to decimal
        d = Decimal(s)
        return f"{d:.2f}"
    
    except (InvalidOperation, ValueError) as e:
        logger.warning(f"Amount parsing error: {amount_str} - {e}")
        return ""


# =========================
# Platform Detection
# =========================
def _detect_platform(row: Dict[str, Any]) -> str:
    """
    ✅ Detect platform from row data
    
    Checks:
    1. U_group field
    2. _route_name metadata
    3. D_vendor_code pattern
    4. _extraction_method metadata
    
    Returns:
        Platform name (META, GOOGLE, etc.)
    """
    try:
        # Check U_group first
        group = row.get("U_group", "").upper()
        if group in PLATFORM_VENDORS:
            return group
        
        # Check metadata
        route = row.get("_route_name", "").lower()
        if "meta" in route:
            return "META"
        if "google" in route:
            return "GOOGLE"
        
        # Check vendor code
        vendor = row.get("D_vendor_code", "").lower()
        if "meta" in vendor or "facebook" in vendor:
            return "META"
        if "google" in vendor:
            return "GOOGLE"
        if "shopee express" in vendor or "spx" in vendor:
            return "SPX"
        if "shopee" in vendor:
            return "SHOPEE"
        if "lazada" in vendor:
            return "LAZADA"
        if "tiktok" in vendor:
            return "TIKTOK"
        
        # Check extraction method
        method = row.get("_extraction_method", "").lower()
        if "meta" in method:
            return "META"
        if "google" in method:
            return "GOOGLE"
        
        return "UNKNOWN"
    
    except Exception as e:
        logger.warning(f"Platform detection error: {e}")
        return "UNKNOWN"


# =========================
# Platform-Specific Validation
# =========================
def _validate_meta_ads(row: Dict[str, Any]) -> List[str]:
    """
    ✅ Validate Meta Ads specific fields
    """
    errors = []
    
    # Must have vendor
    vendor = row.get("D_vendor_code", "")
    if not vendor or "meta" not in vendor.lower():
        errors.append("Meta Ads must have Meta Platforms Ireland as vendor")
    
    # VAT must be NO
    vat = row.get("O_vat_rate", "")
    if vat and vat.upper() != "NO":
        errors.append(f"Meta Ads VAT must be 'NO' (reverse charge), got '{vat}'")
    
    # Must have reference (receipt ID)
    if not row.get("C_reference"):
        errors.append("Meta Ads must have reference (receipt ID)")
    
    # Must have amount
    if not row.get("R_paid_amount"):
        errors.append("Meta Ads must have paid amount")
    
    return errors


def _validate_google_ads(row: Dict[str, Any]) -> List[str]:
    """
    ✅ Validate Google Ads specific fields
    """
    errors = []
    
    # Must have vendor
    vendor = row.get("D_vendor_code", "")
    if not vendor or "google" not in vendor.lower():
        errors.append("Google Ads must have Google Asia Pacific as vendor")
    
    # VAT must be NO
    vat = row.get("O_vat_rate", "")
    if vat and vat.upper() != "NO":
        errors.append(f"Google Ads VAT must be 'NO' (international), got '{vat}'")
    
    # Must have reference (payment number)
    if not row.get("C_reference"):
        errors.append("Google Ads must have reference (payment number)")
    
    # Must have amount
    if not row.get("R_paid_amount"):
        errors.append("Google Ads must have paid amount")
    
    return errors


def _validate_thai_tax(row: Dict[str, Any]) -> List[str]:
    """
    ✅ Validate Thai Tax Invoice fields
    """
    errors = []
    
    # Must have 13-digit tax ID
    tax_id = row.get("E_tax_id_13", "")
    if not tax_id or len(tax_id) != 13:
        errors.append(f"Thai Tax Invoice must have 13-digit tax ID, got '{tax_id}'")
    
    # Must have branch code
    branch = row.get("F_branch_5", "")
    if not branch or len(branch) != 5:
        errors.append(f"Thai Tax Invoice must have 5-digit branch code, got '{branch}'")
    
    # Should have invoice number
    if not row.get("G_invoice_no"):
        errors.append("Thai Tax Invoice should have invoice number")
    
    # Should have VAT (usually 7%)
    if not row.get("O_vat_rate"):
        errors.append("Thai Tax Invoice should have VAT rate")
    
    return errors


def _validate_platform_specific(row: Dict[str, Any], platform: str) -> List[str]:
    """
    ✅ Platform-specific validation dispatcher
    """
    try:
        if platform == "META":
            return _validate_meta_ads(row)
        elif platform == "GOOGLE":
            return _validate_google_ads(row)
        elif platform == "THAI_TAX":
            return _validate_thai_tax(row)
        else:
            return []  # No specific validation
    
    except Exception as e:
        logger.error(f"Platform validation error: {e}")
        return [f"Validation error: {str(e)[:100]}"]


# =========================
# Auto-Correction
# =========================
def _auto_correct_meta_ads(row: Dict[str, Any]) -> Dict[str, Any]:
    """
    ✅ Auto-correct Meta Ads fields
    """
    try:
        # Fix vendor
        vendor = row.get("D_vendor_code", "")
        if not vendor or "meta" not in vendor.lower():
            row["D_vendor_code"] = PLATFORM_VENDORS["META"]
            logger.info("Auto-corrected Meta vendor")
        
        # Fix VAT
        row["O_vat_rate"] = "NO"
        
        # Set U_group
        row["U_group"] = "META"
        
        # Copy metadata to note if available
        method = row.get("_extraction_method")
        if method and not row.get("T_note"):
            row["T_note"] = f"Extraction: {method}"
    
    except Exception as e:
        logger.error(f"Meta auto-correction error: {e}")
    
    return row


def _auto_correct_google_ads(row: Dict[str, Any]) -> Dict[str, Any]:
    """
    ✅ Auto-correct Google Ads fields
    """
    try:
        # Fix vendor
        vendor = row.get("D_vendor_code", "")
        if not vendor or "google" not in vendor.lower():
            row["D_vendor_code"] = PLATFORM_VENDORS["GOOGLE"]
            logger.info("Auto-corrected Google vendor")
        
        # Fix VAT
        row["O_vat_rate"] = "NO"
        
        # Set U_group
        row["U_group"] = "GOOGLE"
        
        # Copy metadata to note
        method = row.get("_extraction_method")
        if method and not row.get("T_note"):
            row["T_note"] = f"Extraction: {method}"
    
    except Exception as e:
        logger.error(f"Google auto-correction error: {e}")
    
    return row


def _auto_correct_platform(row: Dict[str, Any], platform: str) -> Dict[str, Any]:
    """
    ✅ Auto-correct platform-specific fields
    """
    try:
        if platform == "META":
            return _auto_correct_meta_ads(row)
        elif platform == "GOOGLE":
            return _auto_correct_google_ads(row)
        else:
            # Ensure U_group is set
            if not row.get("U_group"):
                row["U_group"] = platform
            return row
    
    except Exception as e:
        logger.error(f"Platform auto-correction error: {e}")
        return row


# =========================
# Helpers
# =========================
def _s(v: Any) -> str:
    """Safe string conversion"""
    if v is None:
        return ""
    try:
        s = str(v).strip()
        if len(s) > MAX_CELL_LENGTH:
            logger.warning(f"Cell value truncated (was {len(s)} chars)")
            s = s[:MAX_CELL_LENGTH]
        return s
    except Exception as e:
        logger.error(f"String conversion error: {e}")
        return ""


def _escape_excel_formula(s: str) -> str:
    """Prevent Excel formula injection"""
    if not s:
        return s
    try:
        if s[0] in EXCEL_INJECTION_PREFIXES:
            return "'" + s
        return s
    except Exception:
        return s


def _compact_no_ws(v: Any) -> str:
    """
    ✅ Remove all whitespace (space/newline/tab)
    """
    s = _s(v)
    if not s:
        return ""
    try:
        return RE_ALL_WS.sub("", s)
    except Exception as e:
        logger.error(f"Compact whitespace error: {e}")
        return s


# =========================
# Preprocessing
# =========================
def _preprocess_rows_for_export(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    ✅ Comprehensive preprocessing with platform awareness
    
    Steps:
    1. Detect platform
    2. Auto-correct platform-specific fields
    3. Renumber sequences
    4. Parse dates to YYYYMMDD
    5. Parse amounts
    6. Compact references
    7. Validate platform-specific rules
    8. Preserve metadata
    """
    out: List[Dict[str, Any]] = []
    seq = 1
    
    logger.info(f"Preprocessing {len(rows)} rows...")
    
    for idx, r in enumerate(rows or [], start=1):
        try:
            rr = dict(r)
            
            # 1. Detect platform
            platform = _detect_platform(rr)
            
            # 2. Auto-correct platform-specific fields
            rr = _auto_correct_platform(rr, platform)
            
            # 3. Sequence number
            rr["A_seq"] = str(seq)
            seq += 1
            
            # 4. Force blank WHT (policy)
            rr["P_wht"] = ""
            
            # 5. Parse dates to YYYYMMDD
            for date_key in DATE_COL_KEYS:
                date_val = rr.get(date_key, "")
                if date_val:
                    parsed = _parse_date_to_yyyymmdd(date_val)
                    rr[date_key] = parsed
                    if not parsed:
                        logger.warning(f"Row {idx}: Could not parse date {date_key}: {date_val}")
            
            # 6. Parse amounts
            for amt_key in ["R_paid_amount", "N_unit_price"]:
                amt_val = rr.get(amt_key, "")
                if amt_val:
                    parsed = _parse_amount(amt_val)
                    rr[amt_key] = parsed
                    if not parsed:
                        logger.warning(f"Row {idx}: Could not parse amount {amt_key}: {amt_val}")
            
            # 7. Compact references (remove spaces/newlines)
            rr["C_reference"] = _compact_no_ws(rr.get("C_reference", ""))
            rr["G_invoice_no"] = _compact_no_ws(rr.get("G_invoice_no", ""))
            
            # 8. Normalize company name
            rr["A_company_name"] = _s(rr.get("A_company_name", ""))
            
            # 9. Ensure critical fields are not None
            for key in ["D_vendor_code", "E_tax_id_13", "F_branch_5"]:
                if rr.get(key) is None:
                    rr[key] = ""
            
            # 10. Validate platform-specific rules
            validation_errors = _validate_platform_specific(rr, platform)
            if validation_errors:
                # Log but don't fail
                logger.warning(f"Row {idx} ({platform}): {'; '.join(validation_errors)}")
                # Add to note
                existing_note = rr.get("T_note", "")
                warning_note = f"⚠️ {'; '.join(validation_errors[:2])}"  # First 2 errors
                rr["T_note"] = f"{existing_note}\n{warning_note}" if existing_note else warning_note
            
            out.append(rr)
        
        except Exception as e:
            logger.error(f"Error preprocessing row {idx}: {e}", exc_info=True)
            continue
    
    logger.info(f"Preprocessing complete: {len(out)} rows ready")
    return out


# =========================
# Validation
# =========================
def validate_rows(rows: List[Dict[str, Any]]) -> Tuple[bool, List[str]]:
    """
    ✅ Validate rows before export
    """
    errors = []
    
    if not rows:
        errors.append("No rows to export")
        return (False, errors)
    
    if len(rows) > MAX_ROWS:
        errors.append(f"Too many rows: {len(rows)} (max: {MAX_ROWS})")
        return (False, errors)
    
    # Check first 100 rows
    for idx, row in enumerate(rows[:100], start=1):
        if not isinstance(row, dict):
            errors.append(f"Row {idx}: Not a dictionary")
            continue
        
        # Check if row has any data
        if not any(row.get(k) for k, _ in COLUMNS):
            errors.append(f"Row {idx}: All fields empty")
    
    if errors:
        return (False, errors)
    
    return (True, [])


# =========================
# Type Conversion
# =========================
def _to_number_or_text(key: str, raw: Any) -> Tuple[Any, str]:
    """
    ✅ Convert value to appropriate Excel type
    """
    try:
        s = _s(raw)
        
        # Always text
        if key in TEXT_COL_KEYS or key in DATE_COL_KEYS:
            return (_escape_excel_formula(s), numbers.FORMAT_TEXT)
        
        # Numeric fields
        if key in NUM_COL_KEYS:
            if key == "M_qty":
                parsed = _parse_amount(s)
                if parsed:
                    try:
                        f = float(parsed)
                        if abs(f - int(f)) < 1e-9:
                            return (int(f), "0")
                        return (f, numbers.FORMAT_NUMBER_00)
                    except Exception:
                        pass
                return (_escape_excel_formula(s), numbers.FORMAT_TEXT)
            
            # Money fields
            parsed = _parse_amount(s)
            if parsed:
                try:
                    return (float(parsed), numbers.FORMAT_NUMBER_00)
                except Exception:
                    pass
            return (_escape_excel_formula(s), numbers.FORMAT_TEXT)
        
        # Default: text
        return (_escape_excel_formula(s), numbers.FORMAT_TEXT)
    
    except Exception as e:
        logger.error(f"Type conversion error for {key}: {e}")
        return ("", numbers.FORMAT_TEXT)


# =========================
# Auto-fit Columns
# =========================
def _auto_fit_columns(ws, max_width: int = 60, min_width: int = 10) -> None:
    """Auto-fit column widths"""
    try:
        for col_idx, (_key, label) in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(label))
            
            max_check = min(ws.max_row + 1, 102)
            for row_idx in range(2, max_check):
                try:
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v is None:
                        continue
                    s = str(v)
                    if "\n" in s:
                        s = s.split("\n", 1)[0]
                    max_len = max(max_len, len(s))
                except Exception:
                    continue
            
            width = int(min(max(max_len + 2, min_width), max_width))
            ws.column_dimensions[col_letter].width = width
    
    except Exception as e:
        logger.error(f"Auto-fit columns error: {e}")


# =========================
# CSV Export
# =========================
def export_rows_to_csv_bytes(rows: List[Dict[str, Any]]) -> bytes:
    """
    ✅ Export rows to CSV with UTF-8 BOM
    """
    try:
        logger.info(f"Starting CSV export for {len(rows)} rows")
        
        # Validate
        is_valid, errors = validate_rows(rows)
        if not is_valid:
            error_msg = "; ".join(errors)
            logger.error(f"Validation failed: {error_msg}")
            raise ExportValidationError(error_msg)
        
        # Preprocess
        rows2 = _preprocess_rows_for_export(rows)
        
        if not rows2:
            raise ExportValidationError("No valid rows after preprocessing")
        
        # Export
        out = io.StringIO()
        wri = csv.writer(out, quoting=csv.QUOTE_MINIMAL)
        
        # Header
        wri.writerow([label for _, label in COLUMNS])
        
        # Data rows
        for r in rows2:
            row_out: List[str] = []
            for k, _label in COLUMNS:
                s = _s(r.get(k, ""))
                s = _escape_excel_formula(s)
                row_out.append(s)
            wri.writerow(row_out)
        
        result = out.getvalue().encode("utf-8-sig")
        logger.info(f"✅ CSV export complete: {len(result)} bytes")
        return result
    
    except ExportValidationError:
        raise
    except Exception as e:
        logger.error(f"CSV export error: {e}", exc_info=True)
        raise Exception(f"CSV export failed: {str(e)}")


# =========================
# XLSX Export
# =========================
def export_rows_to_xlsx_bytes(rows: List[Dict[str, Any]]) -> bytes:
    """
    ✅ Export rows to XLSX with formatting
    """
    try:
        logger.info(f"Starting XLSX export for {len(rows)} rows")
        
        # Validate
        is_valid, errors = validate_rows(rows)
        if not is_valid:
            error_msg = "; ".join(errors)
            logger.error(f"Validation failed: {error_msg}")
            raise ExportValidationError(error_msg)
        
        # Preprocess
        rows2 = _preprocess_rows_for_export(rows)
        
        if not rows2:
            raise ExportValidationError("No valid rows after preprocessing")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "PEAK_IMPORT"
        
        # Header row
        headers = [label for _, label in COLUMNS]
        ws.append(headers)
        
        # Header styling
        header_fill = PatternFill("solid", fgColor="E8F1FF")
        header_font = Font(bold=True)
        header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
        
        thin = Side(style="thin", color="D0D7E2")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        for col_idx in range(1, len(COLUMNS) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = header_align
            c.border = border
        
        # Data rows
        for row_num, r in enumerate(rows2, start=2):
            try:
                values: List[Any] = []
                formats: List[str] = []
                
                for k, _label in COLUMNS:
                    v, fmt = _to_number_or_text(k, r.get(k, ""))
                    values.append(v)
                    formats.append(fmt)
                
                ws.append(values)
                
                current_row = ws.max_row
                for col_idx, fmt in enumerate(formats, start=1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    if fmt:
                        cell.number_format = fmt
                    # Wrap long text
                    cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in {13, 21}))
                    cell.border = border
            
            except Exception as e:
                logger.error(f"Error writing row {row_num}: {e}")
                continue
        
        # Freeze panes and filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
        
        # Force TEXT formatting for critical columns
        col_index = {k: i + 1 for i, (k, _) in enumerate(COLUMNS)}
        for key in (TEXT_COL_KEYS | DATE_COL_KEYS):
            ci = col_index.get(key)
            if not ci:
                continue
            for row_i in range(2, 2 + len(rows2)):
                try:
                    cell = ws.cell(row=row_i, column=ci)
                    cell.number_format = numbers.FORMAT_TEXT
                except Exception:
                    continue
        
        # Auto-fit columns
        _auto_fit_columns(ws)
        
        # Save to bytes
        bio = io.BytesIO()
        wb.save(bio)
        result = bio.getvalue()
        
        logger.info(f"✅ XLSX export complete: {len(result)} bytes")
        return result
    
    except ExportValidationError:
        raise
    except Exception as e:
        logger.error(f"XLSX export error: {e}", exc_info=True)
        raise Exception(f"XLSX export failed: {str(e)}")


# =========================
# Summary
# =========================
def get_export_summary(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    ✅ Get export summary with platform breakdown
    """
    try:
        summary = {
            "total_rows": len(rows),
            "valid_rows": 0,
            "platforms": {},
            "extraction_methods": {},
            "clients": {},
            "date_range": {"earliest": None, "latest": None},
            "total_amount": 0.0,
            "warnings": [],
        }
        
        for row in rows:
            # Count valid rows
            if row.get("D_vendor_code"):
                summary["valid_rows"] += 1
            
            # Detect platform
            platform = _detect_platform(row)
            summary["platforms"][platform] = summary["platforms"].get(platform, 0) + 1
            
            # Extraction methods
            method = row.get("_extraction_method", "unknown")
            summary["extraction_methods"][method] = summary["extraction_methods"].get(method, 0) + 1
            
            # Clients
            company = row.get("A_company_name", "Unknown")
            summary["clients"][company] = summary["clients"].get(company, 0) + 1
            
            # Date range
            doc_date = row.get("B_doc_date")
            if doc_date:
                if not summary["date_range"]["earliest"] or doc_date < summary["date_range"]["earliest"]:
                    summary["date_range"]["earliest"] = doc_date
                if not summary["date_range"]["latest"] or doc_date > summary["date_range"]["latest"]:
                    summary["date_range"]["latest"] = doc_date
            
            # Total amount
            try:
                amt_str = row.get("R_paid_amount", "")
                if amt_str:
                    amt = float(_parse_amount(amt_str))
                    summary["total_amount"] += amt
            except Exception:
                pass
        
        return summary
    
    except Exception as e:
        logger.error(f"Error getting export summary: {e}")
        return {"error": str(e)}


__all__ = [
    "COLUMNS",
    "export_rows_to_csv_bytes",
    "export_rows_to_xlsx_bytes",
    "ExportValidationError",
    "PlatformValidationError",
    "validate_rows",
    "get_export_summary",
]