# -*- coding: utf-8 -*-
# backend/app/services/export_service.py
"""
Export Service - Enhanced v4.3 (POST-PROCESS ALIGNED + SAFE EXPORT + REF NORMALIZER)

✅ Fixes per your request:
1) A_company_name / O_vat_rate / P_wht ต้อง "ออกไปใน XLSX" เสมอ (มี header + มีค่าเท่าที่มี)
2) ห้ามล้าง P_wht ทิ้งใน export (เดิมบังคับ rr["P_wht"] = "" ทำให้ไม่ขึ้น)
3) C_reference / G_invoice_no:
   - compact no-ws เหมือนเดิม
   - ✅ เพิ่ม normalize reference core: ถ้ามี Shopee-TIV-xxxx.pdf -> ตัดให้เหลือ TRS....
   - ✅ sync C == G เสมอ
   - (ถ้า row มี filename แฝง เช่น _filename/filename/source_file จะใช้มันได้)
4) ไม่ยัดข้อความเข้า T_note (policy: ต้องว่าง)
5) Sanitize + validate แบบไม่ทำลายข้อมูล + Excel injection prevention
6) XLSX: auto-fit, freeze, filter, TEXT columns, number format

หมายเหตุ:
- “finalize_row” ต้องอยู่ใน backend/app/extractors/common.py (คุณกำลังแก้ไฟล์นั้นอยู่)
- ไฟล์นี้แก้เรื่อง “Export presentation” และ “normalize reference core” เพิ่มเติม
"""

from __future__ import annotations

import csv
import io
import os
import re
import logging
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Any, Tuple, Optional, Set

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# =========================
# Platform Constants (aligned with your extractors/post_process)
# =========================
PLATFORM_VENDORS = {
    "META": "Meta Platforms Ireland",
    "GOOGLE": "Google Asia Pacific",
    "SHOPEE": "Shopee",
    "LAZADA": "Lazada",
    "TIKTOK": "TikTok",
    "SPX": "Shopee Express",
    "THAI_TAX": "",   # variable
    "UNKNOWN": "Other",
}

# NOTE: price_type values vary by your pipeline. We keep as reference only.
PLATFORM_VAT_RULES = {
    "META": {"J_price_type": "3", "O_vat_rate": "NO"},
    "GOOGLE": {"J_price_type": "3", "O_vat_rate": "NO"},
    "SHOPEE": {"J_price_type": "1", "O_vat_rate": "7%"},
    "LAZADA": {"J_price_type": "1", "O_vat_rate": "7%"},
    "TIKTOK": {"J_price_type": "1", "O_vat_rate": "7%"},
    "SPX": {"J_price_type": "1", "O_vat_rate": "7%"},
    "THAI_TAX": {"J_price_type": "1", "O_vat_rate": "7%"},
    "UNKNOWN": {"J_price_type": "1", "O_vat_rate": "7%"},
}

PLATFORM_GROUPS = {
    "META": "Advertising Expense",
    "GOOGLE": "Advertising Expense",
    "SHOPEE": "Marketplace Expense",
    "LAZADA": "Marketplace Expense",
    "TIKTOK": "Marketplace Expense",
    "SPX": "Delivery/Logistics Expense",
    "THAI_TAX": "General Expense",
    "UNKNOWN": "Other Expense",
}

# =========================
# Columns
# =========================
COLUMNS: List[Tuple[str, str]] = [
    ("A_seq", "ลำดับที่*"),
    ("A_company_name", "ชื่อบริษัท"),
    ("B_doc_date", "วันที่เอกสาร"),
    ("C_reference", "อ้างอิงถึง"),
    ("D_vendor_code", "ผู้รับเงิน/คู่ค้า"),
    ("E_tax_id_13", "เลขทะเบียน 13 หลัก"),
    ("F_branch_5", "เลขสาขา 5 หลัก"),
    ("G_invoice_no", "เลขที่ใบกำกับฯ (ถ้ามี)"),
    ("H_invoice_date", "วันที่ใบกำกับฯ (ถ้ามี)"),
    ("I_tax_purchase_date", "วันที่บันทึกภาษีซื้อ (ถ้ามี)"),
    ("J_price_type", "ประเภทราคา"),
    ("K_account", "บัญชี"),
    ("L_description", "คำอธิบาย"),
    ("M_qty", "จำนวน"),
    ("N_unit_price", "ราคาต่อหน่วย"),
    ("O_vat_rate", "อัตราภาษี"),
    ("P_wht", "หัก ณ ที่จ่าย (ถ้ามี)"),
    ("Q_payment_method", "ชำระโดย"),
    ("R_paid_amount", "จำนวนเงินที่ชำระ"),
    ("S_pnd", "ภ.ง.ด. (ถ้ามี)"),
    ("T_note", "หมายเหตุ"),
    ("U_group", "กลุ่มจัดประเภท"),
]

TEXT_COL_KEYS: Set[str] = {
    "A_seq",
    "A_company_name",
    "C_reference",
    "D_vendor_code",
    "E_tax_id_13",
    "F_branch_5",
    "G_invoice_no",
    "J_price_type",
    "O_vat_rate",
    "S_pnd",
    "Q_payment_method",
    "K_account",
    "P_wht",
}
NUM_COL_KEYS: Set[str] = {"M_qty", "N_unit_price", "R_paid_amount"}
DATE_COL_KEYS: Set[str] = {"B_doc_date", "H_invoice_date", "I_tax_purchase_date"}

# Excel injection prevention
EXCEL_INJECTION_PREFIXES = ("=", "+", "-", "@")

# Regex patterns
RE_YYYYMMDD = re.compile(r"^\d{8}$")
RE_YYYY_MM_DD = re.compile(r"^(\d{4})-(\d{1,2})-(\d{1,2})$")
RE_DD_MM_YYYY = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")
RE_YYYY_SLASH_MM_DD = re.compile(r"^(\d{4})/(\d{1,2})/(\d{1,2})$")
RE_EURO_DD_MM_YYYY = re.compile(r"^(\d{1,2})-(\d{1,2})-(\d{4})$")

RE_ALL_WS = re.compile(r"\s+")
MAX_ROWS = 50000
MAX_CELL_LENGTH = 32767

# =========================
# ✅ Reference core normalizer (ตัด Shopee-TIV- / .pdf / ดึง TRS... ถ้ามี)
# =========================
RE_TRS_CORE = re.compile(r"(TRS[A-Z0-9\-_/.]{10,})", re.IGNORECASE)
RE_RCS_CORE = re.compile(r"(RCS[A-Z0-9\-_/.]{10,})", re.IGNORECASE)
RE_TTSTH_CORE = re.compile(r"(TTSTH\d{10,})", re.IGNORECASE)
RE_LAZ_CORE = re.compile(r"(THMPTI\d{16}|(?:LAZ|LZD)[A-Z0-9\-_/.]{6,}|INV[A-Z0-9\-_/.]{6,})", re.IGNORECASE)

RE_LEADING_NOISE_PREFIX = re.compile(
    r"^(?:Shopee-)?TI[VR]-|^Shopee-|^TIV-|^TIR-|^SPX-|^LAZ-|^LZD-|^TikTok-",
    re.IGNORECASE,
)

def _strip_ext(s: str) -> str:
    return re.sub(r"\.(pdf|png|jpg|jpeg|xlsx|xls)$", "", s, flags=re.IGNORECASE).strip()

def _compact_no_ws(v: Any) -> str:
    s = _s(v)
    if not s:
        return ""
    try:
        return RE_ALL_WS.sub("", s)
    except Exception:
        return s

def _normalize_reference_core(value: Any) -> str:
    """
    Normalize C_reference/G_invoice_no ให้เป็นแกนเอกสารที่ถูกต้อง
    Examples:
      "Shopee-TIV-TRSPEMKP00-00000-251203-0012589.pdf" -> "TRSPEMKP00-00000-251203-0012589"
      "TIV-TRSPEM...." -> "TRSPEM...."
      "TRSPEM...." -> "TRSPEM...."
    """
    s = _compact_no_ws(value)
    if not s:
        return ""
    s = _strip_ext(s)

    # ถ้ามี core ฝังอยู่ ให้ดึง core นั้น
    for pat in (RE_TRS_CORE, RE_RCS_CORE, RE_TTSTH_CORE, RE_LAZ_CORE):
        m = pat.search(s)
        if m:
            return _compact_no_ws(m.group(1))

    # ไม่งั้นตัด prefix noise
    s2 = RE_LEADING_NOISE_PREFIX.sub("", s).strip()
    return _compact_no_ws(s2) if s2 else _compact_no_ws(s)

def _try_get_source_filename(rr: Dict[str, Any]) -> str:
    """
    ถ้า pipeline ใส่ชื่อไฟล์มาด้วย ให้ใช้ช่วย normalize reference ได้แม่นขึ้น
    (ไม่บังคับ แต่รองรับ)
    """
    for k in ("_filename", "filename", "source_file", "_source_file", "_file", "file"):
        v = rr.get(k)
        if v:
            try:
                return os.path.basename(str(v))
            except Exception:
                return str(v)
    return ""

# =========================
# Validation
# =========================
class ExportValidationError(Exception):
    """Raised when export data validation fails"""
    pass

def validate_rows(rows: List[Dict[str, Any]]) -> Tuple[bool, List[str]]:
    errors: List[str] = []
    if not rows:
        return (False, ["No rows to export"])
    if len(rows) > MAX_ROWS:
        return (False, [f"Too many rows: {len(rows)} (max: {MAX_ROWS})"])

    for idx, row in enumerate(rows[:200], start=1):
        if not isinstance(row, dict):
            errors.append(f"Row {idx}: Not a dict")
            continue
        any_value = False
        for k, _ in COLUMNS:
            if row.get(k) not in (None, "", []):
                any_value = True
                break
        if not any_value:
            errors.append(f"Row {idx}: All fields empty")

    if errors:
        return (False, errors)
    return (True, [])

# =========================
# Helpers
# =========================
def _s(v: Any) -> str:
    if v is None:
        return ""
    try:
        s = str(v)
        if len(s) > MAX_CELL_LENGTH:
            logger.warning("Cell value truncated (was %s chars)", len(s))
            s = s[:MAX_CELL_LENGTH]
        return s.strip()
    except Exception:
        return ""

def _escape_excel_formula(s: str) -> str:
    if not s:
        return s
    try:
        if s[:1] in EXCEL_INJECTION_PREFIXES:
            return "'" + s
        return s
    except Exception:
        return s

def _parse_date_to_yyyymmdd(date_str: Any) -> str:
    if not date_str:
        return ""
    s = _s(date_str)

    if RE_YYYYMMDD.match(s):
        return s

    m = RE_YYYY_MM_DD.match(s)
    if m:
        yyyy = m.group(1)
        mm = m.group(2).zfill(2)
        dd = m.group(3).zfill(2)
        return f"{yyyy}{mm}{dd}"

    m = RE_DD_MM_YYYY.match(s)
    if m:
        dd = m.group(1).zfill(2)
        mm = m.group(2).zfill(2)
        yyyy = m.group(3)
        return f"{yyyy}{mm}{dd}"

    m = RE_YYYY_SLASH_MM_DD.match(s)
    if m:
        yyyy = m.group(1)
        mm = m.group(2).zfill(2)
        dd = m.group(3).zfill(2)
        return f"{yyyy}{mm}{dd}"

    m = RE_EURO_DD_MM_YYYY.match(s)
    if m:
        dd = m.group(1).zfill(2)
        mm = m.group(2).zfill(2)
        yyyy = m.group(3)
        return f"{yyyy}{mm}{dd}"

    try:
        from datetime import datetime
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime("%Y%m%d")
            except ValueError:
                continue
    except Exception:
        pass

    return ""

def _parse_amount(amount_str: Any) -> str:
    if amount_str is None:
        return ""
    s = _s(amount_str)
    if not s:
        return ""

    s2 = s.upper()
    s2 = s2.replace("฿", "").replace("THB", "").replace("บาท", "")
    s2 = s2.replace(",", "").replace(" ", "")

    neg = False
    if s2.startswith("(") and s2.endswith(")"):
        neg = True
        s2 = s2[1:-1]

    if s2.startswith("+"):
        s2 = s2[1:]

    if not s2:
        return ""

    try:
        d = Decimal(s2)
        if neg:
            d = -d
        if d < 0:
            return ""
        return f"{d:.2f}"
    except (InvalidOperation, ValueError):
        return ""

def _sync_amount_fields(rr: Dict[str, Any]) -> None:
    n_raw = rr.get("N_unit_price", "")
    r_raw = rr.get("R_paid_amount", "")

    n = _parse_amount(n_raw)
    r = _parse_amount(r_raw)

    if not n and r:
        n = r
    if not r and n:
        r = n
    if n and r and n != r:
        n = r

    rr["N_unit_price"] = n or "0.00"
    rr["R_paid_amount"] = r or "0.00"

def _normalize_qty(rr: Dict[str, Any]) -> None:
    q = _s(rr.get("M_qty", "")) or ""
    if not q:
        rr["M_qty"] = "1"
        return

    q2 = q.replace(",", "").strip()
    qn = _parse_amount(q2)
    if not qn:
        rr["M_qty"] = q2
        return

    try:
        f = float(qn)
        if abs(f - int(f)) < 1e-9:
            rr["M_qty"] = str(int(f))
        else:
            rr["M_qty"] = f"{f:.2f}"
    except Exception:
        rr["M_qty"] = q2

# =========================
# Platform detection (non-destructive)
# =========================
def _detect_platform(rr: Dict[str, Any]) -> str:
    try:
        route = _s(rr.get("_route_name", "")).upper()
        if route in PLATFORM_VENDORS:
            return route

        group = _s(rr.get("U_group", "")).lower()
        vendor = _s(rr.get("D_vendor_code", "")).lower()

        if "advertising" in group or "ads" in group:
            if "meta" in vendor or "facebook" in vendor:
                return "META"
            if "google" in vendor:
                return "GOOGLE"
            if "tiktok" in vendor:
                return "TIKTOK"
            return "UNKNOWN"

        if "marketplace" in group:
            if "shopee" in vendor:
                return "SHOPEE"
            if "lazada" in vendor:
                return "LAZADA"
            if "tiktok" in vendor:
                return "TIKTOK"
            if "spx" in vendor or "express" in vendor:
                return "SPX"

        if "meta" in vendor or "facebook" in vendor:
            return "META"
        if "google" in vendor:
            return "GOOGLE"
        if "lazada" in vendor:
            return "LAZADA"
        if "shopee" in vendor:
            return "SHOPEE"
        if "tiktok" in vendor:
            return "TIKTOK"
        if "spx" in vendor or "express" in vendor:
            return "SPX"

        tax_id = _s(rr.get("E_tax_id_13", ""))
        if len(tax_id) == 13 and tax_id.isdigit():
            return "THAI_TAX"

        return "UNKNOWN"
    except Exception:
        return "UNKNOWN"

# =========================
# Validation warnings (non-destructive)
# =========================
def _add_warning(rr: Dict[str, Any], msg: str) -> None:
    try:
        arr = rr.get("_validation_warnings")
        if not isinstance(arr, list):
            arr = []
        if msg not in arr:
            arr.append(msg)
        rr["_validation_warnings"] = arr[:20]
    except Exception:
        pass

def _validate_platform(rr: Dict[str, Any], platform: str) -> None:
    try:
        if not _s(rr.get("C_reference", "")):
            _add_warning(rr, f"{platform}: missing C_reference")
        if not _s(rr.get("G_invoice_no", "")):
            _add_warning(rr, f"{platform}: missing G_invoice_no")
        if _s(rr.get("T_note", "")):
            _add_warning(rr, f"{platform}: T_note should be empty by policy")

        if platform in ("META", "GOOGLE"):
            if _s(rr.get("O_vat_rate", "")).upper() != "NO":
                _add_warning(rr, f"{platform}: O_vat_rate should be NO")
        else:
            if platform in ("SHOPEE", "LAZADA", "TIKTOK", "SPX") and _s(rr.get("O_vat_rate", "")) not in ("7%", "7", "7.0%", "7.00%"):
                _add_warning(rr, f"{platform}: O_vat_rate expected 7%")

        if _parse_amount(rr.get("R_paid_amount", "")) in ("",):
            _add_warning(rr, f"{platform}: missing/invalid R_paid_amount")
    except Exception:
        return

# =========================
# ✅ Preprocess pipeline
# =========================
def _preprocess_rows_for_export(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    IMPORTANT POLICY:
    - Export must not destroy extractor outputs.
    - Do not write T_note.
    - Keep P_wht if provided (rate-only like "3%") — DO NOT blank it.
    - Normalize references to correct core (TRS... etc) + sync C/G.
    """
    out: List[Dict[str, Any]] = []
    seq = 1

    for idx, r in enumerate(rows or [], start=1):
        try:
            rr = dict(r) if isinstance(r, dict) else {}
            rr["A_seq"] = str(seq)
            seq += 1

            # policy: T_note must be empty
            rr["T_note"] = ""

            # dates normalize (TEXT YYYYMMDD)
            for dk in DATE_COL_KEYS:
                if dk in rr:
                    rr[dk] = _parse_date_to_yyyymmdd(rr.get(dk))

            # ---- Reference normalize (สำคัญ) ----
            # 1) ถ้ามีชื่อไฟล์แฝง ให้ normalize จากชื่อไฟล์นั้นก่อน (แม่นสุด)
            source_file = _try_get_source_filename(rr)
            if source_file:
                ref0 = _normalize_reference_core(source_file)
            else:
                ref0 = ""

            # 2) normalize จากค่าที่มีใน row
            c0 = _normalize_reference_core(rr.get("C_reference", ""))
            g0 = _normalize_reference_core(rr.get("G_invoice_no", ""))

            # 3) เลือก best: filename-ref > C > G
            best_ref = ref0 or c0 or g0

            rr["C_reference"] = best_ref
            rr["G_invoice_no"] = best_ref

            # amounts normalize + sync (ไม่ยุ่งกับ desc/account)
            _sync_amount_fields(rr)

            # qty normalize
            _normalize_qty(rr)

            # ---- Keep P_wht (DON'T BLANK) ----
            # sanitize only (limit length + strip)
            rr["P_wht"] = _s(rr.get("P_wht", ""))

            # ensure group exists
            if not _s(rr.get("U_group", "")):
                platform_guess = _detect_platform(rr)
                rr["U_group"] = PLATFORM_GROUPS.get(platform_guess, rr.get("U_group") or "Marketplace Expense")

            # detect platform + apply VAT defaults only if empty (non-destructive)
            platform = _detect_platform(rr)
            rr["_platform"] = platform

            vat_rule = PLATFORM_VAT_RULES.get(platform, PLATFORM_VAT_RULES["UNKNOWN"])
            if not _s(rr.get("J_price_type", "")):
                rr["J_price_type"] = vat_rule.get("J_price_type", rr.get("J_price_type") or "")
            if not _s(rr.get("O_vat_rate", "")):
                rr["O_vat_rate"] = vat_rule.get("O_vat_rate", rr.get("O_vat_rate") or "")

            # warnings
            _validate_platform(rr, platform)

            # ensure critical columns exist (✅ include requested columns)
            rr["A_company_name"] = _s(rr.get("A_company_name", ""))
            for k in (
                "D_vendor_code", "E_tax_id_13", "F_branch_5",
                "J_price_type", "K_account", "L_description",
                "O_vat_rate", "P_wht", "Q_payment_method", "S_pnd",
            ):
                if rr.get(k) is None:
                    rr[k] = ""

            out.append(rr)

        except Exception as e:
            logger.error("Error preprocessing row %s: %s", idx, e, exc_info=True)
            continue

    return out

# =========================
# Excel type conversion
# =========================
def _to_number_or_text(key: str, raw: Any) -> Tuple[Any, str]:
    s = _s(raw)

    if key in TEXT_COL_KEYS or key in DATE_COL_KEYS:
        return (_escape_excel_formula(s), numbers.FORMAT_TEXT)

    if key in NUM_COL_KEYS:
        if key == "M_qty":
            q = _s(raw)
            if not q:
                return (1, "0")
            qn = _parse_amount(q.replace(",", ""))
            if not qn:
                return (_escape_excel_formula(q), numbers.FORMAT_TEXT)
            try:
                f = float(qn)
                if abs(f - int(f)) < 1e-9:
                    return (int(f), "0")
                return (f, numbers.FORMAT_NUMBER_00)
            except Exception:
                return (_escape_excel_formula(q), numbers.FORMAT_TEXT)

        norm = _parse_amount(s)
        if norm:
            try:
                return (float(norm), numbers.FORMAT_NUMBER_00)
            except Exception:
                return (_escape_excel_formula(s), numbers.FORMAT_TEXT)

        return (_escape_excel_formula(s), numbers.FORMAT_TEXT)

    return (_escape_excel_formula(s), numbers.FORMAT_TEXT)

def _auto_fit_columns(ws, max_width: int = 60, min_width: int = 10) -> None:
    try:
        for col_idx, (_key, label) in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(label))

            max_check = min(ws.max_row + 1, 220)
            for row_idx in range(2, max_check):
                try:
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v is None:
                        continue
                    s = str(v)
                    if "\n" in s:
                        s = s.split("\n", 1)[0]
                    max_len = max(max_len, len(s))
                except Exception:
                    continue

            ws.column_dimensions[col_letter].width = int(min(max(max_len + 2, min_width), max_width))
    except Exception as e:
        logger.error("Auto-fit columns error: %s", e)

# =========================
# CSV Export
# =========================
def export_rows_to_csv_bytes(rows: List[Dict[str, Any]]) -> bytes:
    try:
        is_valid, errors = validate_rows(rows)
        if not is_valid:
            raise ExportValidationError("; ".join(errors))

        rows2 = _preprocess_rows_for_export(rows)
        if not rows2:
            raise ExportValidationError("No valid rows after preprocessing")

        out = io.StringIO()
        wri = csv.writer(out, quoting=csv.QUOTE_MINIMAL)

        wri.writerow([label for _, label in COLUMNS])

        for r in rows2:
            row_out: List[str] = []
            for k, _label in COLUMNS:
                s = _s(r.get(k, ""))
                s = _escape_excel_formula(s)
                row_out.append(s)
            wri.writerow(row_out)

        return out.getvalue().encode("utf-8-sig")

    except ExportValidationError:
        raise
    except Exception as e:
        logger.error("CSV export error: %s", e, exc_info=True)
        raise Exception(f"CSV export failed: {str(e)}")

# =========================
# XLSX Export
# =========================
def export_rows_to_xlsx_bytes(rows: List[Dict[str, Any]]) -> bytes:
    try:
        is_valid, errors = validate_rows(rows)
        if not is_valid:
            raise ExportValidationError("; ".join(errors))

        rows2 = _preprocess_rows_for_export(rows)
        if not rows2:
            raise ExportValidationError("No valid rows after preprocessing")

        wb = Workbook()
        ws = wb.active
        ws.title = "PEAK_IMPORT"

        ws.append([label for _, label in COLUMNS])

        header_fill = PatternFill("solid", fgColor="E8F1FF")
        header_font = Font(bold=True)
        header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)

        thin = Side(style="thin", color="D0D7E2")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx in range(1, len(COLUMNS) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = header_align
            c.border = border

        for r in rows2:
            values: List[Any] = []
            formats: List[str] = []
            for k, _label in COLUMNS:
                v, fmt = _to_number_or_text(k, r.get(k, ""))
                values.append(v)
                formats.append(fmt)

            ws.append(values)
            row_i = ws.max_row
            for col_idx, fmt in enumerate(formats, start=1):
                cell = ws.cell(row=row_i, column=col_idx)
                if fmt:
                    cell.number_format = fmt
                cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in {13, 21}))
                cell.border = border

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

        col_index = {k: i + 1 for i, (k, _) in enumerate(COLUMNS)}
        for key in (TEXT_COL_KEYS | DATE_COL_KEYS):
            ci = col_index.get(key)
            if not ci:
                continue
            for row_i in range(2, 2 + len(rows2)):
                try:
                    ws.cell(row=row_i, column=ci).number_format = numbers.FORMAT_TEXT
                except Exception:
                    continue

        _auto_fit_columns(ws)

        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()

    except ExportValidationError:
        raise
    except Exception as e:
        logger.error("XLSX export error: %s", e, exc_info=True)
        raise Exception(f"XLSX export failed: {str(e)}")

# =========================
# Summary
# =========================
def get_export_summary(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    try:
        summary: Dict[str, Any] = {
            "total_rows": len(rows or []),
            "valid_rows": 0,
            "by_platform": {},
            "by_group": {},
            "by_vendor": {},
            "clients": {},
            "date_range": {"earliest": None, "latest": None},
            "total_amount": 0.0,
            "warnings": [],
        }

        warn_acc: List[str] = []

        for rr0 in rows or []:
            rr = rr0 if isinstance(rr0, dict) else {}
            if _s(rr.get("D_vendor_code", "")):
                summary["valid_rows"] += 1

            platform = _detect_platform(rr)
            summary["by_platform"][platform] = summary["by_platform"].get(platform, 0) + 1

            group = _s(rr.get("U_group", "Unknown")) or "Unknown"
            summary["by_group"][group] = summary["by_group"].get(group, 0) + 1

            vendor = _s(rr.get("D_vendor_code", "Unknown")) or "Unknown"
            summary["by_vendor"][vendor] = summary["by_vendor"].get(vendor, 0) + 1

            company = _s(rr.get("A_company_name", "Unknown")) or "Unknown"
            summary["clients"][company] = summary["clients"].get(company, 0) + 1

            doc_date = _parse_date_to_yyyymmdd(rr.get("B_doc_date"))
            if doc_date:
                if not summary["date_range"]["earliest"] or doc_date < summary["date_range"]["earliest"]:
                    summary["date_range"]["earliest"] = doc_date
                if not summary["date_range"]["latest"] or doc_date > summary["date_range"]["latest"]:
                    summary["date_range"]["latest"] = doc_date

            amt = _parse_amount(rr.get("R_paid_amount", ""))
            if amt:
                try:
                    summary["total_amount"] += float(amt)
                except Exception:
                    pass

            vw = rr.get("_validation_warnings")
            if isinstance(vw, list):
                for w in vw[:3]:
                    if isinstance(w, str):
                        warn_acc.append(w)

        seen = set()
        uniq = []
        for w in warn_acc:
            if w not in seen:
                uniq.append(w)
                seen.add(w)
        summary["warnings"] = uniq[:10]
        return summary

    except Exception as e:
        logger.error("Error getting export summary: %s", e, exc_info=True)
        return {"error": str(e)}

__all__ = [
    "COLUMNS",
    "PLATFORM_VENDORS",
    "PLATFORM_VAT_RULES",
    "PLATFORM_GROUPS",
    "export_rows_to_csv_bytes",
    "export_rows_to_xlsx_bytes",
    "ExportValidationError",
    "validate_rows",
    "get_export_summary",
]
